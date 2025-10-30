#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Gera arquivo de teste Status com dados
"""

from openpyxl import Workbook
from openpyxl.styles import Font

# Criar workbook
wb = Workbook()
ws = wb.active

if ws is not None:
    ws.title = 'Status'
    
    # Headers (como estão no Excel original)
    headers = ['NUMERO', 'ETAPA', 'PRAZO', 'SLA HORAS', 'TEMPO', 'ENTROU', 'USUÁRIO', 'SAIU', 'USUÁRIO.1', 'MOVIMENTAÇÃO', 'TAG ATIVIDADE']
    
    # Escrever headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if cell:
            cell.font = Font(bold=True)
    
    # Dados de teste - múltiplas movimentações por NUMERO
    dados = [
        ('24323348', 'ARQUIVADO', 'DENTRO PRAZO', '0', '28740:56:33', '2021-01-25 17:03:35', 'MARIANA SOARES MARINS', '2025-08-28 06:00:08', 'Usuário Master', '', '#PREDRETONAR'),
        ('24323348', 'ENVIADO', 'DENTRO PRAZO', '24', '100:12:45', '2021-01-20 10:00:00', 'JOÃO SILVA', '2021-01-25 17:03:35', 'MARIANA SOARES MARINS', 'Movimentado', '#PREDRETONAR'),
        ('38254050', 'ENVIADO BKO', 'DENTRO PRAZO', '48', '2025-10-30 00:19:12', '2023-02-15 10:44:12', 'TATIANE KELLY', '2023-02-15 11:03:24', 'JOAO PAULO', '', '#FASTCHIP,#MOVEL'),
        ('38254050', 'AGUARDANDO RETORNO', 'DENTRO PRAZO', '48', '2025-10-30 05:36:39', '2023-02-15 11:41:20', 'ALINE SILVA', '2023-02-15 17:17:59', 'ALINE SILVA', '', '#FASTCHIP,#MOVEL'),
    ]
    
    # Repetir 25 vezes para ter 100 linhas
    dados_expandidos = dados * 25
    
    # Escrever dados
    for row_idx, record in enumerate(dados_expandidos, 2):
        for col_idx, valor in enumerate(record, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 19
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 19
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30
    
    # Salvar
    output_file = 'downloads/Exportacao Status.xlsx'
    wb.save(output_file)
    
    print(f'✅ Arquivo criado: {output_file}')
    print(f'   Linhas: {len(dados_expandidos)} registros de teste')
    print(f'   NUMEROs únicos: 2 (24323348, 38254050) com múltiplas movimentações cada')
else:
    print('❌ Erro ao criar worksheet')
